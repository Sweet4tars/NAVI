from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .case_studies import get_case_study


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D96535")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)


def export_case_study_excel(case_id: str, output_path: str | Path) -> Path:
    case = get_case_study(case_id)
    return export_share_payload_excel(case, output_path)


def export_share_payload_excel(case: dict, output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览"
    _fill_overview(overview, case)
    _fill_route(workbook.create_sheet("路线"), case)
    _fill_days(workbook.create_sheet("每日行程"), case)
    _fill_stays(workbook.create_sheet("住宿位置"), case)
    _fill_meals(workbook.create_sheet("餐饮候选"), case)
    _fill_budget(workbook.create_sheet("预算"), case)

    for sheet in workbook.worksheets:
        _autosize(sheet)

    workbook.save(output)
    return output


def _fill_overview(sheet, case: dict) -> None:
    sheet["A1"] = case["title"]
    sheet["A1"].font = TITLE_FONT
    rows = [
        ("副标题", case["subtitle"]),
        ("日期", case["date_range"]),
        ("出行方式", case["transport_mode"]),
        ("同行规模", case["travelers"]),
        ("预算目标", case["budget_target"]),
        ("路线", " -> ".join(case["route_nodes"])),
        ("摘要", case["summary"]),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        sheet[f"A{row_index}"] = label
        sheet[f"B{row_index}"] = value
    _style_header_row(sheet, 2, ["字段", "内容"])


def _fill_route(sheet, case: dict) -> None:
    headers = ["起点", "终点", "距离", "车程", "强度", "说明"]
    _style_header_row(sheet, 1, headers)
    for row_index, leg in enumerate(case["route_legs"], start=2):
        sheet.cell(row=row_index, column=1, value=leg["from"])
        sheet.cell(row=row_index, column=2, value=leg["to"])
        sheet.cell(row=row_index, column=3, value=leg["distance"])
        sheet.cell(row=row_index, column=4, value=leg["drive_time"])
        sheet.cell(row=row_index, column=5, value=leg["intensity"])
        sheet.cell(row=row_index, column=6, value=leg["note"])


def _fill_days(sheet, case: dict) -> None:
    headers = ["Day", "日期", "标题", "强度", "落脚城市", "优先住区", "行程安排"]
    _style_header_row(sheet, 1, headers)
    for row_index, day in enumerate(case["days"], start=2):
        schedule = "\n".join(day["schedule"])
        values = [
            day["day_index"],
            day["date"],
            day["title"],
            day["intensity"],
            day["base_city"],
            day["priority_zone"],
            schedule,
        ]
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _fill_stays(sheet, case: dict) -> None:
    headers = ["城市", "优先住区", "次选住区", "位置理由", "酒店/民宿", "类型", "价格", "推荐理由", "来源链接"]
    _style_header_row(sheet, 1, headers)
    row_index = 2
    for stay in case["stay_recommendations"]:
        for hotel in stay["hotel_candidates"]:
            values = [
                stay["city"],
                stay["priority_zone"],
                stay["secondary_zone"],
                stay["zone_reason"],
                hotel["name"],
                hotel["type"],
                hotel["price"],
                hotel["reason"],
                hotel["source_url"],
            ]
            for col_index, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row_index += 1


def _fill_meals(sheet, case: dict) -> None:
    headers = ["Day", "日期", "城市", "餐段", "候选", "来源", "推荐理由", "来源链接"]
    _style_header_row(sheet, 1, headers)
    row_index = 2
    for day in case["days"]:
        for meal in day["meals"]:
            for candidate in meal["candidates"]:
                values = [
                    day["day_index"],
                    day["date"],
                    day["base_city"],
                    meal["slot"],
                    candidate["name"],
                    candidate["source"],
                    candidate["reason"],
                    candidate["url"],
                ]
                for col_index, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_index, column=col_index, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                row_index += 1


def _fill_budget(sheet, case: dict) -> None:
    _style_header_row(sheet, 1, ["项", "金额"])
    budget = case["budget"]
    rows = [
        ("酒店", budget["hotel_total"]),
        ("车费", budget["car_total"]),
        ("餐饮", budget["meal_total"]),
        ("总计", budget["grand_total"]),
    ]
    for row_index, (label, value) in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)


def _style_header_row(sheet, row_index: int, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            value = value.split("\n", 1)[0]
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 44)
